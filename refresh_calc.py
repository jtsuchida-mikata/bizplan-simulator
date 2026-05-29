#!/usr/bin/env python3
"""xlsm の「計算結果」シートを読み、レバー列を文字列に変換して
   計算結果_729.xlsx に上書き保存する。convert.py の入力形式に揃える。"""
import openpyxl
from datetime import datetime

SRC = "/Users/jummatsuchida/Desktop/エンジニアのミカタ/ファイナンス資料/LP作成/新事業計画_25 (20260529)_共有.xlsm"
DST = "/Users/jummatsuchida/Desktop/エンジニアのミカタ/ファイナンス資料/LP作成/計算結果_729.xlsx"

# 月次ヘッダーが日付型(datetime)で来る版に対応。convert.py はシリアル値(整数)前提なので
# Excelシリアル値へ正規化する（年度/四半期は文字列なのでそのまま）。
_EPOCH = datetime(1899, 12, 30)
def _norm_hdr(v):
    if isinstance(v, datetime):
        return (v - _EPOCH).days
    return v

LEVER = {1: "楽観", 2: "標準", 3: "悲観"}
# 旧ヘッダーに揃える（col0-7）
OLD_HDR_HEAD = ["combo_id", "PP入社数", "PP離職率", "粗利率",
                "営業成約件数", "採用単価", "管理部門人数", "KPI項目"]

print(f"読み込み: {SRC}")
src = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
src_ws = src["計算結果"]
it = src_ws.iter_rows(values_only=True)

src_hdr = list(next(it))
print(f"  新ヘッダー列数: {len(src_hdr)}")

dst = openpyxl.Workbook(write_only=True)
dst_ws = dst.create_sheet("計算結果")
# col0-7 は旧名に統一、col8以降（期間列）は新ヘッダーを流用（月次の日付はシリアル値へ）
new_hdr = OLD_HDR_HEAD + [_norm_hdr(v) for v in src_hdr[8:]]
dst_ws.append(new_hdr)

NCOL = len(new_hdr)
n_rows = 0
for row in it:
    r = list(row)
    for i in range(1, 7):
        v = r[i]
        if isinstance(v, int):
            r[i] = LEVER.get(v, v)
    # 末尾の空セルが切り詰められている行を110列にパディング
    if len(r) < NCOL:
        r += [""] * (NCOL - len(r))
    dst_ws.append(r)
    n_rows += 1

print(f"  データ行数: {n_rows}")
print(f"書き出し: {DST}")
dst.save(DST)
import os
print(f"完了。サイズ: {os.path.getsize(DST)/1_000_000:.1f} MB")
